import os
import time
import pickle
import logging
import openpyxl
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

load_dotenv()

CONFIG = {"ARQUIVO_ALVO": "Setor Cancelamento.xlsx", "SETOR_ALVO": "Setor Cancelamento"}

NOVA_SENHA = "Mudar@1243!!"

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")


def configurar_driver(headless=False):
    options = Options()
    if headless:
        options.add_argument("--headless")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--disable-gpu")

    driver = webdriver.Chrome(options=options)
    if not headless:
        driver.maximize_window()
    return driver


def salvar_cookies(driver, arquivo="cookies.pkl"):
    with open(arquivo, "wb") as f:
        pickle.dump(driver.get_cookies(), f)
    logging.info("Cookies salvos com sucesso!")


def carregar_cookies(driver, arquivo="cookies.pkl"):
    if os.path.exists(arquivo):
        try:
            logging.info("Carregando cookies salvos...")
            driver.get("https://directinternet.hubsoft.com.br")
            with open(arquivo, "rb") as f:
                cookies = pickle.load(f)
                for cookie in cookies:
                    driver.add_cookie(cookie)
            driver.refresh()
            return True
        except Exception:
            return False
    return False


def esperar_e_clicar(wait, driver, by, value, nome_elemento):
    try:
        logging.info(f"--- Clicando em: {nome_elemento}")
        element = wait.until(EC.element_to_be_clickable((by, value)))
        element.click()
        return element
    except Exception:
        logging.warning(f"Clique normal falhou em {nome_elemento}. Forçando via JS...")
        try:
            element = driver.find_element(by, value)
            driver.execute_script("arguments[0].click();", element)
            return element
        except Exception as e:
            raise Exception(f"Não foi possível interagir com {nome_elemento}: {e}")


def realizar_login(driver, wait, email, senha):
    if carregar_cookies(driver):
        try:
            wait.until(EC.url_contains("/dashboard"))
            logging.info("Login via Cookies realizado!")
            return
        except TimeoutException:
            logging.warning("Cookies expirados.")

    try:
        logging.info("Realizando login manual...")
        driver.get("https://directinternet.hubsoft.com.br/login")
        esperar_e_clicar(
            wait,
            driver,
            By.XPATH,
            "//input[@type='email' or contains(@name, 'mail')]",
            "Email",
        ).send_keys(email)
        esperar_e_clicar(
            wait, driver, By.XPATH, "//button[contains(., 'Validar')]", "Validar"
        )
        esperar_e_clicar(
            wait, driver, By.XPATH, "//input[@type='password']", "Senha"
        ).send_keys(senha)
        esperar_e_clicar(
            wait, driver, By.XPATH, "//button[contains(., 'Entrar')]", "Entrar"
        )

        logging.info("\n--- REALIZE O 2FA NO NAVEGADOR AGORA ---\n")
        wait.until(EC.url_contains("/dashboard"))
        salvar_cookies(driver)
        logging.info("Login confirmado.")
    except Exception as e:
        logging.error(f"Erro login: {e}")
        raise


def alterar_setor_dinamico(driver, wait, action, nome_setor):
    logging.info(f"Alterando setor para: {nome_setor}")
    menu_setores = wait.until(
        lambda d: [
            el
            for el in d.find_elements(By.XPATH, "//md-select[@name='setor']")
            if el.is_displayed()
        ][0]
    )
    menu_setores.click()

    try:
        campo_filtro = wait.until(EC.element_to_be_clickable((By.ID, "input_187")))
    except:
        campo_filtro = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//md-select-menu[contains(@class,'md-active')]//input[@type='search']",
                )
            )
        )

    campo_filtro.send_keys(Keys.CONTROL + "a")
    campo_filtro.send_keys(Keys.BACKSPACE)
    campo_filtro.send_keys(nome_setor)
    time.sleep(1)

    xpath_item = f"//button[contains(@aria-label, '{nome_setor}')]"
    for _ in range(5):
        try:
            if driver.find_element(By.XPATH, xpath_item).is_displayed():
                break
        except:
            action.send_keys(Keys.PAGE_DOWN).perform()
            time.sleep(0.5)

    botao_item = wait.until(EC.presence_of_element_located((By.XPATH, xpath_item)))
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao_item)
    time.sleep(0.5)
    try:
        botao_item.click()
    except:
        driver.execute_script("arguments[0].click();", botao_item)

    time.sleep(0.5)
    action.send_keys(Keys.ESCAPE).perform()


def alterar_permissao_dinamica(driver, wait, nome_permissao):
    logging.info(f"Alterando permissão para: {nome_permissao}")
    xpath_permissoes = "//button[@aria-label='PERMISSÕES']"

    try:
        botao_permissoes = wait.until(
            EC.element_to_be_clickable((By.XPATH, xpath_permissoes))
        )
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", botao_permissoes
        )
        time.sleep(0.5)
        botao_permissoes.click()
    except Exception:
        botao_permissoes = driver.find_element(By.XPATH, xpath_permissoes)
        driver.execute_script("arguments[0].click();", botao_permissoes)

    time.sleep(1)

    menu_permissao = wait.until(
        lambda d: [
            el
            for el in d.find_elements(By.XPATH, "//md-select[@name='tipo_permissao']")
            if el.is_displayed()
        ][0]
    )
    menu_permissao.click()

    grupo_valor = "grupo"
    xpath_opcao_ativa = (
        f"//div[contains(@class, 'md-active')]//md-option[@value='{grupo_valor}']"
    )
    time.sleep(1)

    try:
        opcao_correta = wait.until(
            EC.element_to_be_clickable((By.XPATH, xpath_opcao_ativa))
        )
        opcao_correta.click()
    except Exception:
        opcao_correta = driver.find_element(By.XPATH, xpath_opcao_ativa)
        driver.execute_script("arguments[0].click();", opcao_correta)

    time.sleep(2)

    menu_grupo = wait.until(
        lambda d: [
            el
            for el in d.find_elements(
                By.XPATH, "//md-select[@name='Grupo de Permissão']"
            )
            if el.is_displayed()
        ][0]
    )
    menu_grupo.click()

    try:
        campo_filtro = wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//div[contains(@class, 'md-active')]//input[@type='search']",
                )
            )
        )
        campo_filtro.click()
        campo_filtro.send_keys(Keys.CONTROL + "a")
        campo_filtro.send_keys(Keys.BACKSPACE)
        campo_filtro.send_keys(nome_permissao)

        time.sleep(2)

        xpath_botao = f"//div[contains(@class, 'md-active')]//button[@aria-label='{nome_permissao}']"

        sucesso = False
        for tentativa in range(3):
            try:
                botao_final = wait.until(
                    EC.presence_of_element_located((By.XPATH, xpath_botao))
                )
                driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});", botao_final
                )
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", botao_final)
                sucesso = True
                break
            except Exception:
                time.sleep(1)

        if not sucesso:
            raise Exception("Não foi possível clicar no item após a pesquisa.")

    except Exception as e:
        logging.warning(f"Erro fatal ao selecionar a permissão: {e}")


def processar_usuario(driver, wait, action, nome_usuario):
    try:
        logging.info(f"\nPROCESSANDO: {nome_usuario}")
        driver.get("https://directinternet.hubsoft.com.br/configuracao/geral/usuario")
        time.sleep(2.5)

        campo_busca = wait.until(
            EC.element_to_be_clickable((By.ID, "configuracao-fiscal-search"))
        )
        campo_busca.click()
        campo_busca.send_keys(Keys.CONTROL + "a")
        campo_busca.send_keys(Keys.BACKSPACE)
        campo_busca.send_keys(nome_usuario)
        time.sleep(1)
        campo_busca.send_keys(Keys.ENTER)
        time.sleep(2)

        esperar_e_clicar(
            wait, driver, By.XPATH, "//button[@aria-label='Ações']", "Botão Ações"
        )
        time.sleep(0.5)
        esperar_e_clicar(
            wait, driver, By.XPATH, "//button[@aria-label='Editar']", "Botão Editar"
        )

        logging.info("Alterando senha...")
        campo_senha = wait.until(EC.element_to_be_clickable((By.ID, "senha")))
        campo_senha.click()
        campo_senha.send_keys(Keys.CONTROL + "a")
        campo_senha.send_keys(Keys.BACKSPACE)
        campo_senha.send_keys(NOVA_SENHA)

        campo_confirma = wait.until(
            EC.element_to_be_clickable((By.ID, "confirmar_senha"))
        )
        campo_confirma.click()
        campo_confirma.send_keys(Keys.CONTROL + "a")
        campo_confirma.send_keys(Keys.BACKSPACE)
        campo_confirma.send_keys(NOVA_SENHA)

        alterar_setor_dinamico(driver, wait, action, CONFIG["SETOR_ALVO"])
        alterar_permissao_dinamica(driver, wait, CONFIG["SETOR_ALVO"])

        logging.info("Salvando...")
        botao_salvar = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, "//button[@type='submit' and @ng-click='vm.save()']")
            )
        )
        driver.execute_script("arguments[0].click();", botao_salvar)

        time.sleep(3)
        logging.info(f"Usuário '{nome_usuario}' finalizado!")

    except Exception as e:
        logging.error(f"Erro ao processar {nome_usuario}: {e}")


def ler_nomes_do_excel(caminho_ficheiro):
    lista_nomes = []
    if not os.path.exists(caminho_ficheiro):
        raise FileNotFoundError(f"Ficheiro '{caminho_ficheiro}' não encontrado.")

    workbook = openpyxl.load_workbook(caminho_ficheiro, data_only=True)
    sheet = workbook.active

    coluna_usuario = None
    for cell in sheet[1]:
        if cell.value and str(cell.value).strip().lower() == "usuario":
            coluna_usuario = cell.column
            break

    if not coluna_usuario:
        raise ValueError("Coluna 'usuario' não encontrada na primeira linha do Excel.")

    for row in range(2, sheet.max_row + 1):
        nome = sheet.cell(row=row, column=coluna_usuario).value
        if nome:
            lista_nomes.append(str(nome).strip())

    return lista_nomes


def main():
    EMAIL = os.getenv("HUBSOFT_EMAIL")
    SENHA = os.getenv("HUBSOFT_SENHA")

    logging.info(f"A ler utilizadores do ficheiro: {CONFIG['ARQUIVO_ALVO']}")
    try:
        usuarios = ler_nomes_do_excel(CONFIG["ARQUIVO_ALVO"])
        logging.info(f"Total de utilizadores encontrados: {len(usuarios)}")
    except Exception as e:
        logging.error(f"Erro ao ler ficheiro: {e}")
        return

    driver = configurar_driver(headless=False)
    wait = WebDriverWait(driver, 30)
    action = ActionChains(driver)

    try:
        realizar_login(driver, wait, email=EMAIL, senha=SENHA)

        for nome in usuarios:
            processar_usuario(driver, wait, action, nome)

    finally:
        input("\nPressione Enter para encerrar...")
        driver.quit()


if __name__ == "__main__":
    main()
